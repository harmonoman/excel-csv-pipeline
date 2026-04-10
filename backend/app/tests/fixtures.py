"""
Test fixture helpers for building minimal .xlsx workbooks in memory.
These are NOT the full T7-1 fixture library — they are lightweight
builders used exclusively to support parser unit tests.
"""
import io
import openpyxl
 
 
def build_workbook(sheets: list[dict]) -> io.BytesIO:
    """
    Build an in-memory .xlsx workbook from a list of sheet definitions.
 
    Each sheet definition:
        {
            "name": "Alpha Fund",
            "rows": [
                ["row 1 value", None, None],   # metadata
                ["First", "Last", "Amount"],   # header
                ["John", "Doe", "100.00"],      # data
            ]
        }
 
    Returns a BytesIO object ready for openpyxl.load_workbook().
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
 
    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["name"])
        for row in sheet_def["rows"]:
            ws.append(row)
 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer