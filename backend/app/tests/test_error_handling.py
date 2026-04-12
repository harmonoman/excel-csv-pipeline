"""
Tests for T4-4 — System-level error handling.

Verifies that:
- Corrupted/unreadable xlsx returns HTTP 400 with structured error
- Pipeline stage failures return HTTP 400 with stage context
- Unexpected exceptions return HTTP 500 with safe message (no stack trace)
- Valid file with invalid rows returns HTTP 200 (validation != system failure)
- Response schema is consistent across all error types
"""
import io
from unittest.mock import patch
from fastapi.testclient import TestClient

from app.main import app
from app.processing.pipeline import PipelineError

# Standard client — re-raises server exceptions (good for most tests)
client = TestClient(app)

# Client for testing 500 handlers — lets the app handle exceptions rather than re-raising
client_no_raise = TestClient(app, raise_server_exceptions=False)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_xlsx_upload(content: bytes, filename: str = "test.xlsx"):
    """Helper to create a file upload tuple."""
    return ("file", (filename, io.BytesIO(content), XLSX_MIME))


def make_valid_xlsx() -> bytes:
    """Build a minimal valid xlsx file in memory."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["First", "Last", "Address1", "City", "State", "Zip", "GiftDate", "Amount"])
    ws.append(["John", "Doe", "123 Main St", "Nashville", "TN", "37201", "2024-01-01", "100.00"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Corrupted file → HTTP 400
# ---------------------------------------------------------------------------

def test_corrupted_xlsx_returns_400():
    """Uploading a corrupted (non-parseable) xlsx returns HTTP 400."""
    corrupted = b"this is not an xlsx file at all"
    response = client.post("/upload", files=[make_xlsx_upload(corrupted)])
    assert response.status_code == 400


def test_corrupted_xlsx_returns_structured_error():
    """Corrupted file error response has error.type and error.stage fields."""
    corrupted = b"corrupted bytes"
    response = client.post("/upload", files=[make_xlsx_upload(corrupted)])
    body = response.json()
    assert "error" in body
    assert "type" in body["error"]
    assert "stage" in body["error"]
    assert "message" in body["error"]


def test_corrupted_xlsx_stage_is_parse():
    """Corrupted file error identifies parse as the failing stage."""
    corrupted = b"not xlsx"
    response = client.post("/upload", files=[make_xlsx_upload(corrupted)])
    body = response.json()
    assert body["error"]["stage"] == "parse"


def test_corrupted_xlsx_no_stack_trace():
    """Corrupted file error response must not expose Python stack traces."""
    corrupted = b"not xlsx"
    response = client.post("/upload", files=[make_xlsx_upload(corrupted)])
    body = response.json()
    response_text = str(body)
    assert "Traceback" not in response_text
    assert "File \"" not in response_text


# ---------------------------------------------------------------------------
# Pipeline stage failure → HTTP 400
# ---------------------------------------------------------------------------

def test_pipeline_error_returns_400():
    """PipelineError raised during processing returns HTTP 400."""
    valid_xlsx = make_valid_xlsx()
    with patch("app.main.run_pipeline", side_effect=PipelineError(
        stage="normalize",
        error_type="NormalizationError",
        error_message="normalization failed",
    )):
        response = client.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    assert response.status_code == 400


def test_pipeline_error_response_contains_stage():
    """PipelineError response includes the failing stage name."""
    valid_xlsx = make_valid_xlsx()
    with patch("app.main.run_pipeline", side_effect=PipelineError(
        stage="normalize",
        error_type="NormalizationError",
        error_message="normalization failed",
    )):
        response = client.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    body = response.json()
    assert body["error"]["stage"] == "normalize"
    assert body["error"]["type"] == "NormalizationError"


# ---------------------------------------------------------------------------
# Unexpected exception → HTTP 500
# ---------------------------------------------------------------------------

def test_unexpected_exception_returns_500():
    """Unhandled Exception inside pipeline returns HTTP 500."""
    valid_xlsx = make_valid_xlsx()
    with patch("app.main.run_pipeline", side_effect=RuntimeError("something exploded")):
        response = client_no_raise.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    assert response.status_code == 500


def test_unexpected_exception_safe_message():
    """HTTP 500 response does not expose stack trace or internal details."""
    valid_xlsx = make_valid_xlsx()
    with patch("app.main.run_pipeline", side_effect=RuntimeError("something exploded")):
        response = client_no_raise.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    body = response.json()
    assert "error" in body
    assert "Traceback" not in str(body)
    assert "something exploded" not in str(body)  # internal detail must not leak


def test_unexpected_exception_stage_is_unknown():
    """HTTP 500 response uses 'unknown' as stage when exception is unclassified."""
    valid_xlsx = make_valid_xlsx()
    with patch("app.main.run_pipeline", side_effect=RuntimeError("boom")):
        response = client_no_raise.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    body = response.json()
    assert body["error"]["stage"] == "unknown"


# ---------------------------------------------------------------------------
# Validation failures → HTTP 200 (not a system error)
# ---------------------------------------------------------------------------

def test_valid_file_with_invalid_rows_returns_200():
    """File with rows that fail validation returns HTTP 200 — not an error.
    Also confirms rejected_rows > 0 so we know validation actually ran.

    Note: all-None rows are filtered by the parser as completely empty.
    Use rows with actual invalid values (bad state, bad zip) to test rejection.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["First", "Last", "Address1", "City", "State", "Zip", "GiftDate", "Amount"])
    ws.append(["John", "Doe", "123 Main St", "Nashville", "TN", "37201", "2024-01-01", "100.00"])
    # Invalid row — bad state code, bad zip, negative amount
    ws.append(["Jane", "Smith", "456 Oak Ave", "Memphis", "XX", "1234", "2024-01-02", "-50.00"])
    buf = io.BytesIO()
    wb.save(buf)

    response = client.post("/upload", files=[make_xlsx_upload(buf.getvalue())])
    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["total_rows"] == 2
    assert body["summary"]["rejected_rows"] > 0  # validator ran and caught the bad row
    assert body["summary"]["clean_rows"] == 1


def test_valid_file_with_invalid_rows_has_summary():
    """Successful pipeline response includes summary with row counts."""
    valid_xlsx = make_valid_xlsx()
    response = client.post("/upload", files=[make_xlsx_upload(valid_xlsx)])
    assert response.status_code == 200
    body = response.json()
    assert "summary" in body
    assert "total_rows" in body["summary"]
    assert "clean_rows" in body["summary"]
    assert "rejected_rows" in body["summary"]


def test_validation_failure_does_not_raise_exception():
    """Row-level validation failures must never produce HTTP 4xx or 5xx."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["First", "Last", "Address1", "City", "State", "Zip", "GiftDate", "Amount"])
    # All rows invalid — bad state, bad zip, bad amount
    ws.append(["John", "Doe", "123 Main", "Nashville", "XX", "1234", "not-a-date", "-100"])
    buf = io.BytesIO()
    wb.save(buf)

    response = client.post("/upload", files=[make_xlsx_upload(buf.getvalue())])
    assert response.status_code == 200


# ---------------------------------------------------------------------------
# Edge cases identified in E4 checkpoint review
# ---------------------------------------------------------------------------

def test_empty_workbook_returns_200_with_zero_rows():
    """
    Valid xlsx with no parseable sheets (e.g. Instructions-only workbook)
    returns HTTP 200 with total_rows: 0. This is correct — no pipeline failure
    occurred, the file was simply empty of usable data.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws.append(["This sheet contains setup instructions"])
    ws.append(["Do not delete"])
    buf = io.BytesIO()
    wb.save(buf)

    response = client.post("/upload", files=[make_xlsx_upload(buf.getvalue())])
    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["total_rows"] == 0
    assert body["summary"]["clean_rows"] == 0
    assert body["summary"]["rejected_rows"] == 0