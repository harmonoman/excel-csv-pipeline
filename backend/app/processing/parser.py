"""
parser.py — Multi-sheet Excel workbook parser.
 
Responsibilities:
- Read all sheets dynamically (no hardcoded sheet names)
- Detect header row per sheet using alias matching from mapping.json
- Normalize column names (lowercase, strip whitespace)
- Build a DataFrame per valid sheet with _source_sheet column
- Concatenate all valid sheets into a single DataFrame
 
This module does NOT:
- Map columns to canonical names (T3-1)
- Inject Client field (T3-3)
- Validate or reject rows (T4)
"""
import io
import logging
 
import openpyxl
import pandas as pd
 
logger = logging.getLogger(__name__)
 
HEADER_MATCH_THRESHOLD = 2
 
 
def _build_alias_set(mapping_config: dict) -> set:
    """
    Flatten all aliases from mapping.json into a single lowercase set
    for fast O(1) header row matching.
    """
    aliases = set()
    for field_aliases in mapping_config["fields"].values():
        for alias in field_aliases:
            aliases.add(alias.lower().strip())
    return aliases
 
 
def _detect_header_row(
    rows: list[tuple],
    alias_set: set,
    scan_rows: int,
) -> int | None:
    """
    Scan the first N rows and return the 0-based index of the first row
    with >= HEADER_MATCH_THRESHOLD alias matches.
 
    Accepts pre-read rows (list of tuples) to avoid double-iteration
    issues with openpyxl read_only worksheets.
 
    Returns None if no valid header row found within scan_rows.
    """
    for row_idx, row in enumerate(rows[:scan_rows]):
        normalized = [
            str(cell).lower().strip()
            for cell in row
            if cell is not None
        ]
        matches = sum(1 for val in normalized if val in alias_set)
        if matches >= HEADER_MATCH_THRESHOLD:
            return row_idx
 
    return None
 
 
def _build_dataframe(
    rows: list[tuple],
    header_row_idx: int,
    sheet_name: str,
) -> pd.DataFrame | None:
    """
    Build a DataFrame from pre-read rows starting at the detected header row.
 
    - Normalizes column names (lowercase, stripped)
    - Filters out completely empty data rows
    - Adds _source_sheet column
    - Returns None if no data rows exist after the header
 
    Note: duplicate column names are allowed at this stage —
    deduplication is not the parser's responsibility.
    """
    header = [
        str(val).lower().strip() if val is not None else f"unnamed_{i}"
        for i, val in enumerate(rows[header_row_idx])
    ]
 
    data_rows = rows[header_row_idx + 1:]
 
    if not data_rows:
        return None
 
    # Filter out completely empty rows
    data_rows = [
        row for row in data_rows
        if any(cell is not None for cell in row)
    ]
 
    if not data_rows:
        return None
 
    df = pd.DataFrame(data_rows, columns=header)
    df["_source_sheet"] = sheet_name
    return df
 
 
def parse_workbook(
    source: str | io.BytesIO,
    mapping_config: dict,
) -> pd.DataFrame:
    """
    Parse all sheets in an Excel workbook into a single DataFrame.
 
    Each sheet is read once — rows are materialized into a list before
    any processing to avoid openpyxl read_only generator exhaustion.
 
    Args:
        source: file path string or BytesIO object
        mapping_config: loaded and validated mapping.json config
 
    Returns:
        Single concatenated DataFrame from all valid sheets.
        Empty DataFrame if no valid sheets found.
    """
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    alias_set = _build_alias_set(mapping_config)
    scan_rows = mapping_config["header_scan_rows"]
 
    valid_frames = []
 
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
 
        # Read all rows once — read_only worksheets are generators
        # and can only be iterated once. Materializing here prevents
        # silent empty results from double-iteration.
        all_rows = list(sheet.iter_rows(values_only=True))
 
        header_row_idx = _detect_header_row(all_rows, alias_set, scan_rows)
 
        if header_row_idx is None:
            logger.warning(
                "No valid header found in sheet '%s' — sheet excluded from processing",
                sheet_name,
            )
            continue
 
        df = _build_dataframe(all_rows, header_row_idx, sheet_name)
 
        if df is None:
            # Header found but no data rows — silently skip
            continue
 
        valid_frames.append(df)
 
    wb.close()
 
    if not valid_frames:
        return pd.DataFrame()
 
    return pd.concat(valid_frames, ignore_index=True)
 